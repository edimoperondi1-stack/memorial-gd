"""
step3_gerar_xlsx.py
-------------------
Extrai a aba SAIDA do arquivo preenchido e salva como o Excel de saída final.

Abordagem: copia o workbook inteiro, remove todas as abas exceto SAIDA,
e substitui fórmulas pelos valores calculados (preservando formatação).

Replica exatamente o que a macro VBA faz:
  PasteSpecial xlPasteValues em A1:FA5 → novo workbook com 1 aba → salva como .xlsx
"""

import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.properties import PageSetupProperties


# Intervalo copiado pela macro: A1:FA5
COL_MAX = column_index_from_string("FA")   # 157
ROW_MAX = 5


def _patch_empty_inline_strings(caminho_xlsx: str) -> None:
    """
    Corrige um bug do openpyxl: células com valor "" são serializadas como
    <c t="inlineStr"></c> (sem <is><t/></is>), fazendo load_workbook retornar
    None em vez de "". Este patch reescreve o XML para incluir <is><t/></is>.
    """
    tmp = caminho_xlsx + ".patch.tmp"
    try:
        with zipfile.ZipFile(caminho_xlsx, "r") as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("xl/worksheets/"):
                        xml = data.decode("utf-8")
                        # Transforma <c ... t="inlineStr"></c>  →  <c ... t="inlineStr"><is><t/></is></c>
                        xml = re.sub(
                            r'(<c [^>]*t="inlineStr")[^<]*(/>|></c>)',
                            r'\1><is><t/></is></c>',
                            xml,
                        )
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)
        Path(tmp).replace(caminho_xlsx)
    except Exception as e:
        Path(tmp).unlink(missing_ok=True)
        raise RuntimeError(f"Falha ao aplicar patch de strings vazias: {e}") from e


def _normalizar_valor(val):
    """
    Normaliza tipos para garantir consistência no output.
    - None permanece None
    - Strings são stripped
    - Números ficam como número
    """
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        return v if v else None   # string vazia → None
    return val


def _aplicar_ajustes_formato(ws) -> None:
    """
    Aplica os 8 ajustes de formato para que o XLSX de saída seja idêntico
    à planilha oficial de referência aceita pela Energisa.

    Deve ser chamada APÓS o preenchimento de valores e ANTES de wb.save().
    """
    # Correção 1 — Células AT2:DR2 devem ser string vazia, não None
    # AT=46, DR=122 inclusive
    for col in range(46, 123):
        cell = ws.cell(row=2, column=col)
        if cell.value is None:
            cell.value = ""

    # Correção 2 — F2 (telefone), G2 (CPF), M2 (CEP) devem ser tipo numérico
    for col_letter in ["F", "G", "M"]:
        cell = ws[f"{col_letter}2"]
        if isinstance(cell.value, str):
            cleaned = (
                cell.value
                .replace(".", "")
                .replace("-", "")
                .replace("/", "")
                .replace(" ", "")
            )
            if cleaned.isdigit():
                cell.value = int(cleaned)
            else:
                try:
                    cell.value = float(cleaned)
                except ValueError:
                    pass  # manter como string se não converter

    # Correção 3 — Sheet protection na aba SAIDA (sem senha, como na planilha oficial)
    ws.protection = SheetProtection(sheet=True, password=None)

    # Correção 4 — Aba SAIDA deve ficar oculta (sheet_state = "hidden")
    # Definido aqui; aplicado após remoção das outras abas para evitar conflito.
    ws.sheet_state = "hidden"

    # Correção 5 — Formato numérico de AJ2 deve ser "General"
    ws["AJ2"].number_format = "General"

    # Correção 6 — Largura das colunas DL (116) e DM (117) = 36.43
    ws.column_dimensions[get_column_letter(116)].width = 36.43  # DL
    ws.column_dimensions[get_column_letter(117)].width = 36.43  # DM

    # Correção 7 — Margens de header e footer de impressão = 0.315 polegadas
    ws.page_margins.header = 0.315
    ws.page_margins.footer = 0.315

    # Correção 8 — Remover fitToWidth e fitToHeight da configuração de página
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None
    # Garantir que fitToPage não está ativo
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = False


def gerar_xlsx(caminho_preenchido: str, pasta_saida: str, nome_titular: str, codigo_uc: str) -> str:
    """
    Gera o Excel de saída final, preservando a formatação original da aba SAIDA.

    Abordagem:
      1. Copia o arquivo inteiro (preserva tudo)
      2. Lê os valores calculados (data_only=True) separadamente
      3. No arquivo copiado, substitui fórmulas por valores na aba SAIDA
      4. Remove todas as abas exceto SAIDA
      5. Salva com o nome correto

    Args:
        caminho_preenchido: arquivo .xlsx com fórmulas já recalculadas.
        pasta_saida: pasta onde salvar o output.
        nome_titular: usado no nome do arquivo.
        codigo_uc: usado no nome do arquivo.

    Returns:
        Caminho absoluto do arquivo .xlsx gerado.
    """
    # 1. Ler os VALORES calculados da aba SAIDA (LibreOffice já recalculou)
    wb_vals = load_workbook(caminho_preenchido, data_only=True)
    if "SAIDA" not in wb_vals.sheetnames:
        raise ValueError("Aba 'SAIDA' não encontrada no arquivo preenchido.")

    ws_vals = wb_vals["SAIDA"]
    valores = {}
    for row in ws_vals.iter_rows(min_row=1, max_row=ROW_MAX, max_col=COL_MAX):
        for cell in row:
            if cell.value is not None:
                valores[(cell.row, cell.column)] = _normalizar_valor(cell.value)
    wb_vals.close()

    # 2. Copiar o arquivo inteiro para temp (preserva formatação, cores, etc.)
    tmp_path = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(caminho_preenchido, tmp_path)

    # 3. Abrir o arquivo copiado (com fórmulas + formatação intacta)
    wb = load_workbook(tmp_path)
    ws = wb["SAIDA"]

    # 4. Substituir fórmulas por valores na aba SAIDA (linhas 1-5, colunas A-FA)
    for row in range(1, ROW_MAX + 1):
        for col in range(1, COL_MAX + 1):
            cell = ws.cell(row=row, column=col)
            # Se tinha fórmula, substituir pelo valor calculado
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                val = valores.get((row, col))
                cell.value = val
            # Se não é fórmula mas o valor calculado existe, usar ele
            elif (row, col) in valores:
                cell.value = valores[(row, col)]

    # 5. Remover todas as abas exceto SAIDA
    # Tornar SAIDA visível temporariamente para permitir a remoção das outras abas
    ws.sheet_state = "visible"
    for nome_aba in list(wb.sheetnames):
        if nome_aba != "SAIDA":
            del wb[nome_aba]

    # 5b. Criar Planilha1 como aba auxiliar visível (obrigatória para poder ocultar SAIDA)
    # A planilha oficial de referência também possui esta aba
    wb.create_sheet("Planilha1")

    # 5c. Aplicar os 8 ajustes de formato (inclui proteção e sheet_state=hidden)
    _aplicar_ajustes_formato(ws)

    # 5d. Garantir que SAIDA permanece como aba ativa APÓS ocultá-la
    # (deve ser definido depois de sheet_state=hidden para não ser revertido)
    wb.active = wb.index(ws)

    # 6. Salvar com o nome correto
    nome_titular_sanitizado = nome_titular.upper().strip()
    nome_arquivo = f"{nome_titular_sanitizado}_UC_{codigo_uc}.xlsx"

    pasta = Path(pasta_saida)
    pasta.mkdir(parents=True, exist_ok=True)
    caminho_saida = str(pasta / nome_arquivo)

    wb.save(caminho_saida)
    wb.close()

    # 6b. Patch pós-salvamento: corrige células "" serializadas incorretamente pelo openpyxl
    _patch_empty_inline_strings(caminho_saida)

    # Limpar temp
    Path(tmp_path).unlink(missing_ok=True)

    print(f"  [step3] OK — Excel de saida gerado: {nome_arquivo}")
    return caminho_saida


def validar_xlsx_saida(caminho_xlsx: str) -> dict:
    """
    Valida o Excel gerado verificando campos obrigatórios na linha 2.
    Retorna dict com os valores dos campos críticos e lista de problemas.
    """
    wb = load_workbook(caminho_xlsx, data_only=True)
    ws = wb["SAIDA"] if "SAIDA" in wb.sheetnames else wb.active

    # Mapear cabeçalhos da linha 1
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value)

    # Ler valores da linha 2
    dados = {}
    for cell in ws[2]:
        header = headers.get(cell.column, f"Col{cell.column}")
        dados[header] = cell.value

    wb.close()

    # Campos obrigatórios que não podem ser None
    obrigatorios = ["UC", "Cliente", "Logradouro:", "Cidade:", "UF:", "Potencia geração"]
    problemas = [f"'{c}' está vazio" for c in obrigatorios if not dados.get(c)]

    return {
        "campos": dados,
        "problemas": problemas,
        "ok": len(problemas) == 0,
    }
