"""
step2_recalcular.py
-------------------
Recalcula todas as fórmulas do Excel preenchido usando LibreOffice headless.
Verifica erros críticos no resultado e retorna um relatório.
"""

import json
import os
import subprocess
import sys
from pathlib import Path

# Reutiliza o script de recalc do skill xlsx
# Procura primeiro em lo_scripts/ (Docker), depois no skills dir (Cowork)
_POSSIBLE_RECALC = [
    Path(__file__).parent.parent / "lo_scripts" / "recalc.py",
    Path("/app/lo_scripts/recalc.py"),
    Path("/sessions/intelligent-adoring-hamilton/mnt/.claude/skills/xlsx/scripts/recalc.py"),
]
RECALC_SCRIPT = next((p for p in _POSSIBLE_RECALC if p.exists()), _POSSIBLE_RECALC[0])
SKILLS_DIR = str(RECALC_SCRIPT.parent)

# Campos críticos que NÃO devem ter erro após recalculo
CAMPOS_CRITICOS = [
    "CONFIG!D2",   # UCF consolidado
    "CONFIG!E2",   # TITULAR consolidado
    "CONFIG!AM2",  # Empresa distribuidora
    "SAIDA!A2",    # UC
    "SAIDA!B2",    # Cliente
]

# Erros aceitáveis (existem na planilha original e não afetam o output)
# Podem ser células exatas ("CONFIG!Q27") ou prefixos de aba ("INVERSOR-MODULO!")
ERROS_ACEITAVEIS_CELULAS = {"CONFIG!Q27"}

# Abas inteiras onde erros #N/A são aceitáveis (VLOOKUPs de base de equipamentos)
ERROS_ACEITAVEIS_ABAS = {"INVERSOR-MODULO"}


def recalcular(caminho_xlsx: str, timeout: int = 90) -> dict:
    """
    Recalcula as fórmulas do arquivo via LibreOffice.

    Args:
        caminho_xlsx: caminho do arquivo .xlsx preenchido.
        timeout: segundos máximos para aguardar o LibreOffice.

    Returns:
        dict com status, erros encontrados e resumo.
    """
    if not Path(caminho_xlsx).exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_xlsx}")

    # Inserir o diretório do skills no path para recalc.py funcionar
    if SKILLS_DIR not in sys.path:
        sys.path.insert(0, SKILLS_DIR)

    print(f"  [step2] Recalculando fórmulas: {Path(caminho_xlsx).name} ...")

    resultado_raw = subprocess.run(
        [sys.executable, str(RECALC_SCRIPT), caminho_xlsx, str(timeout)],
        capture_output=True,
        text=True,
        timeout=timeout + 30,
    )

    if resultado_raw.returncode != 0 and not resultado_raw.stdout.strip():
        raise RuntimeError(
            f"LibreOffice falhou (código {resultado_raw.returncode}):\n"
            f"{resultado_raw.stderr[:500]}"
        )

    try:
        resultado = json.loads(resultado_raw.stdout.strip())
    except json.JSONDecodeError:
        raise RuntimeError(
            f"Saída inesperada do recalc:\n{resultado_raw.stdout[:300]}"
        )

    # Filtrar erros aceitáveis
    erros_reais = {}
    if resultado.get("status") == "errors_found":
        for tipo_erro, info in resultado.get("error_summary", {}).items():
            locais_reais = []
            for loc in info.get("locations", []):
                # Pular células explicitamente aceitas
                if loc in ERROS_ACEITAVEIS_CELULAS:
                    continue
                # Pular abas inteiras (ex: INVERSOR-MODULO!A2 → aba = INVERSOR-MODULO)
                aba = loc.split("!")[0] if "!" in loc else ""
                if aba in ERROS_ACEITAVEIS_ABAS:
                    continue
                locais_reais.append(loc)
            if locais_reais:
                erros_reais[tipo_erro] = {"count": len(locais_reais), "locations": locais_reais}

    status_final = "ok" if not erros_reais else "erros_criticos"

    relatorio = {
        "status": status_final,
        "total_formulas": resultado.get("total_formulas", 0),
        "erros_ignorados": list(ERROS_ACEITAVEIS_CELULAS) + [f"{a}!*" for a in ERROS_ACEITAVEIS_ABAS],
        "erros_criticos": erros_reais,
    }

    if status_final == "ok":
        print(f"  [step2] OK — Recalculo OK — {relatorio['total_formulas']} fórmulas calculadas")
    else:
        print(f"  [step2] ERRO — Erros criticos encontrados: {erros_reais}")

    return relatorio


def verificar_campos_criticos(caminho_xlsx: str) -> dict:
    """
    Após recalcular, lê os campos críticos e verifica se têm valores válidos.
    Retorna dict com campo → valor.
    """
    from openpyxl import load_workbook

    wb = load_workbook(caminho_xlsx, data_only=True)
    resultado = {}

    mapa = {
        "CONFIG!D2":  ("CONFIG",  "D2"),
        "CONFIG!E2":  ("CONFIG",  "E2"),
        "CONFIG!AM2": ("CONFIG",  "AM2"),
        "CONFIG!K20": ("CONFIG",  "K20"),  # código empresa (ex: EMT)
        "SAIDA!A2":   ("SAIDA",   "A2"),
        "SAIDA!B2":   ("SAIDA",   "B2"),
        "SAIDA!AJ2":  ("SAIDA",   "AJ2"),  # potência de geração final
    }

    problemas = []
    for label, (aba, cel) in mapa.items():
        if aba in wb.sheetnames:
            val = wb[aba][cel].value
            resultado[label] = val
            if val is None or val == "" or val == 0:
                problemas.append(f"{label} = {val!r}  ← VAZIO ou ZERO")

    wb.close()

    if problemas:
        print(f"  [step2] AVISO — Campos com atencao apos recalculo:")
        for p in problemas:
            print(f"           {p}")
    else:
        print(f"  [step2] OK — Todos os campos críticos têm valores.")

    resultado["_problemas"] = problemas
    return resultado
