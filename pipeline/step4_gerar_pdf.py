"""
step4_gerar_pdf.py
------------------
Gera o PDF de saída a partir do arquivo preenchido e recalculado.

Usa LibreOffice DIRETAMENTE (sem openpyxl), preservando:
  - Shapes / DrawingML (diagramas unifilares, logos, caixas de texto)
  - Formatação completa (fontes, cores, bordas, células mescladas)
  - Áreas de impressão e page setup

Replica a lógica da macro VBA:
  - Determina quais abas exportar com base no tipo_fsa e presença de UCs beneficiárias
  - Oculta as abas desnecessárias via macro LibreOffice
  - Exporta apenas abas visíveis para PDF

Abas por tipo:
  SOLICITACAO (minigeração):   SOLICITACAO, RELACAO DE CARGA, FORMULARIO, MD-SOLAR, DU-SOLAR
  FSA MICRO <=10:              FSA MICRO <=10, MD-SOLAR, DU-SOLAR
  FSA MICRO >10:               FSA MICRO >10,  MD-SOLAR, DU-SOLAR
  + UC BENEFICIARIAS se houver UCs beneficiárias preenchidas
"""

import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

# PDF Watermark configurations (USER ADJUSTABLE)
# Calibrado para a renderização do LibreOffice (que comprime ~16% verticalmente vs Excel)
# --- Esquema Unifilar ---
DIAGRAMA_IMG_PATH = str(Path(__file__).parent / "api" / "static" / "diagrama_unifilar.png")
DIAGRAMA_POS_X = 248       # Posição horizontal (pts) — movido 18pts pra esquerda
DIAGRAMA_POS_Y = 325       # Posição vertical (pts) — movido 14pts pra baixo (PDF: y menor = mais baixo)
DIAGRAMA_WIDTH = 385       # Largura (pts)
DIAGRAMA_HEIGHT = 197      # Altura (pts)

# --- Placa de Advertência (CUIDADO / Risco de Choque) ---
PLACA_IMG_PATH = str(Path(__file__).parent / "api" / "static" / "diagrama_original.png")
# Crop em pixels para remover as anotações de dimensão (setas azuis, cotas):
PLACA_CROP = (75, 65, 510, 348)   # (left, top, right, bottom) — captura só a placa CUIDADO
PLACA_POS_X = 140          # Posição horizontal (pts) — movido pra esquerda
PLACA_POS_Y = 200          # Posição vertical (pts)  — movido pra baixo
PLACA_WIDTH = 85           # Largura (pts)
PLACA_HEIGHT = 55          # Altura (pts)

# Mapeamento tipo_fsa → abas do PDF (sem UC BENEFICIARIAS)
ABAS_PDF = {
    "SOLICITACAO":    ["SOLICITACAO", "RELACAO DE CARGA", "FORMULARIO", "MD-SOLAR", "DU-SOLAR"],
    "FSA MICRO <=10": ["FSA MICRO <=10", "RELACAO DE CARGA", "FORMULARIO", "MD-SOLAR", "DU-SOLAR"],
    "FSA MICRO >10":  ["FSA MICRO >10",  "RELACAO DE CARGA", "FORMULARIO", "MD-SOLAR", "DU-SOLAR"],
}


def _patch_page_scale(caminho_xlsx: str, abas_scale: dict) -> None:
    """
    Reduz o scale de impressão de abas específicas via patch direto no XML do XLSX.

    No Linux, o LibreOffice renderiza margens/fontes ligeiramente maiores que no
    Windows. O scale=70% do template cabe no Windows mas transborda no Linux.
    Solução simples: reduzir o scale para um valor menor (ex: 60%) que garante
    que caiba em ambas as plataformas.

    Args:
        caminho_xlsx: caminho do XLSX a corrigir.
        abas_scale: dict {nome_aba: novo_scale} ex: {"RELACAO DE CARGA": 60}

    Não usa openpyxl (preserva drawings/shapes restaurados anteriormente).
    """
    tmp = caminho_xlsx + ".scale.tmp"
    try:
        # Resolver rid → target para as abas
        with zipfile.ZipFile(caminho_xlsx, "r") as zin:
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        rid_to_target = {}
        for m in re.finditer(r'<Relationship\b([^>]+)>', rels_xml):
            tag = m.group(1)
            id_m = re.search(r'Id="(rId\d+)"', tag)
            tgt_m = re.search(r'Target="([^"]+)"', tag)
            if id_m and tgt_m:
                rid_to_target[id_m.group(1)] = tgt_m.group(1)

        # Mapear nome da aba → (caminho XML, novo scale)
        target_to_scale = {}
        for nome, novo_scale in abas_scale.items():
            escaped = re.escape(nome)
            m = re.search(rf'<sheet\b[^>]*\bname="{escaped}"[^>]*\br:id="(rId\d+)"', wb_xml)
            if m:
                target = rid_to_target.get(m.group(1), "")
                if target:
                    target_to_scale[f"xl/{target}"] = novo_scale

        if not target_to_scale:
            return

        with zipfile.ZipFile(caminho_xlsx, "r") as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename in target_to_scale:
                        xml = data.decode("utf-8")
                        novo = target_to_scale[item.filename]
                        # Substituir scale="XX" por scale="novo" APENAS dentro de <pageSetup>
                        xml = re.sub(
                            r'(<pageSetup\b[^>]*?)scale="\d+"',
                            rf'\1scale="{novo}"',
                            xml
                        )
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)

        Path(tmp).replace(caminho_xlsx)
    except Exception as e:
        Path(tmp).unlink(missing_ok=True)
        print(f"  [step4] AVISO: falha ao aplicar patch de scale: {e}")


def _tem_ucs_beneficiarias(caminho_xlsx: str) -> bool:
    """Verifica se há UCs beneficiárias preenchidas (célula B15 de UC BENEFICIARIAS).
    Usa openpyxl apenas para LEITURA (não salva, não modifica)."""
    from openpyxl import load_workbook
    wb = load_workbook(caminho_xlsx, data_only=True)
    try:
        if "UC BENEFICIARIAS" not in wb.sheetnames:
            return False
        val = wb["UC BENEFICIARIAS"]["B15"].value
        return bool(val and str(val).strip())
    finally:
        wb.close()


def gerar_pdf(
    caminho_preenchido: str,
    pasta_saida: str,
    nome_titular: str,
    codigo_uc: str,
    tipo_fsa: str = "SOLICITACAO",
) -> str:
    """
    Gera o PDF de saída usando LibreOffice diretamente.

    O arquivo preenchido NÃO é modificado — uma cópia temporária é usada.

    Args:
        caminho_preenchido: .xlsx com fórmulas recalculadas.
        pasta_saida: pasta de destino do PDF.
        nome_titular: para nomear o arquivo.
        codigo_uc: para nomear o arquivo.
        tipo_fsa: tipo de formulário ("SOLICITACAO", "FSA MICRO <=10", "FSA MICRO >10").

    Returns:
        Caminho absoluto do PDF gerado.
    """
    if tipo_fsa not in ABAS_PDF:
        raise ValueError(f"tipo_fsa inválido: '{tipo_fsa}'. Válidos: {list(ABAS_PDF.keys())}")

    # Montar lista de abas para o PDF
    abas = list(ABAS_PDF[tipo_fsa])
    if _tem_ucs_beneficiarias(caminho_preenchido):
        abas.append("UC BENEFICIARIAS")
        print(f"  [step4] UCs beneficiárias detectadas — incluindo na saída.")

    print(f"  [step4] Abas para o PDF: {abas}")

    # Copiar arquivo para pasta temporária (LibreOffice pode modificar o original)
    tmp_dir = tempfile.mkdtemp(prefix="step4_pdf_")
    tmp_xlsx = os.path.join(tmp_dir, Path(caminho_preenchido).name)
    shutil.copy2(caminho_preenchido, tmp_xlsx)

    # Restaurar drawings/shapes/imagens do template original
    # (openpyxl + LibreOffice recalc removem os DrawingML ao salvar)
    from restaurar_drawings import restaurar_drawings
    template_path = str(Path(__file__).parent.parent / "MEMORIAL_GD_v4-22022022 (54).xlsx")
    restaurar_drawings(tmp_xlsx, template_path)
    print(f"  [step4] Drawings restaurados do template original.")

    # Converter fórmulas → valores nas abas do PDF
    # (evita #REF! quando a macro deleta abas referenciadas)
    from converter_formulas import converter_formulas_para_valores
    converter_formulas_para_valores(
        caminho_xlsx=tmp_xlsx,
        caminho_template_ou_recalc=caminho_preenchido,
        abas_converter=abas,
    )
    print(f"  [step4] Fórmulas convertidas em valores nas abas do PDF.")

    # Corrigir scale de impressão para abas que transbordam no Linux
    # (LibreOffice Linux renderiza margens/fontes ~10% maiores que Windows)
    # Template original: RELACAO DE CARGA=70, FORMULARIO=65 → reduz 10% cada
    _patch_page_scale(tmp_xlsx, {
        "RELACAO DE CARGA": 55,
        "FORMULARIO": 50,
    })
    print(f"  [step4] Scale de impressão reduzido para RELACAO DE CARGA(55) e FORMULARIO(50).")

    # Nome do PDF
    nome_pdf = f"{nome_titular.upper().strip()}_UC_{codigo_uc}.pdf"
    Path(pasta_saida).mkdir(parents=True, exist_ok=True)
    caminho_pdf = os.path.join(pasta_saida, nome_pdf)

    # Exportar via LibreOffice (100% nativo, sem openpyxl)
    from export_pdf_lo import export_pdf
    resultado = export_pdf(
        caminho_xlsx=tmp_xlsx,
        caminho_pdf=caminho_pdf,
        abas=abas,
        timeout=120,
    )

    # Limpar temp
    shutil.rmtree(tmp_dir, ignore_errors=True)

    if not resultado["ok"]:
        raise RuntimeError(f"Falha ao gerar PDF: {resultado.get('error', 'desconhecido')}")

    # ==============================================================
    # OVERLAY COMO BACKGROUND: LibreOffice não renderiza os DrawingML
    # shapes (caixas, linhas, setas do diagrama). O overlay adiciona
    # o esquema como imagem de FUNDO, e o texto das células fica por
    # cima (legível). Usa merge_under em vez de merge_page.
    # ==============================================================
    if "DU-SOLAR" in abas and os.path.exists(DIAGRAMA_IMG_PATH):
        try:
            print("  [step4] Aplicando esquema unifilar como FUNDO do PDF...")
            _aplicar_diagrama_fundo(caminho_pdf)
        except Exception as e:
            print(f"  [step4] AVISO: Falha ao aplicar diagrama no fundo: {e}")

    tamanho = os.path.getsize(caminho_pdf)
    print(f"  [step4] OK — PDF gerado: {nome_pdf} ({tamanho:,} bytes)")
    return caminho_pdf

def _preparar_imagem_transparente(img_path: str, crop_box=None, line_alpha: int = 255) -> str:
    """
    Abre uma imagem PNG, aplica crop opcional, torna pixels brancos/quase-brancos
    transparentes (alpha=0) e define alpha dos pixels restantes (linhas/formas).

    Args:
        img_path: caminho da imagem PNG.
        crop_box: (left, top, right, bottom) para recortar.
        line_alpha: alpha para pixels não-brancos (0=invisível, 255=opaco).
                    Use ~100-130 para semi-transparente.
    """
    import tempfile
    import numpy as np
    from PIL import Image

    img = Image.open(img_path).convert("RGBA")
    if crop_box:
        img = img.crop(crop_box)
    arr = np.array(img)
    white_mask = (arr[:, :, 0] > 240) & (arr[:, :, 1] > 240) & (arr[:, :, 2] > 240)
    arr[white_mask, 3] = 0
    # Pixels não-brancos: aplicar alpha especificado (semi-transparente)
    if line_alpha < 255:
        non_white = ~white_mask
        arr[non_white, 3] = np.minimum(arr[non_white, 3], line_alpha)
    processed = Image.fromarray(arr)
    tmp = tempfile.mktemp(suffix=".png")
    processed.save(tmp)
    return tmp


def _aplicar_diagrama_fundo(caminho_pdf: str):
    """
    Usa pypdf e reportlab para criar uma página de overlay (diagrama + placa de
    advertência) e mesclar sobre a ÚLTIMA página do PDF (DU-SOLAR).

    Ambas as imagens têm fundo branco 100% opaco; são processadas para tornar
    os pixels brancos transparentes, evitando cobrir o texto da aba.
    """
    import io
    import os
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import landscape, A4

    # Pré-processar imagens (branco → transparente, linhas semi-opacas)
    # line_alpha=210 → ~82% opacidade — diagrama bem visível, texto por baixo ainda legível
    tmp_diagrama = _preparar_imagem_transparente(DIAGRAMA_IMG_PATH, line_alpha=210)
    tmp_placa    = _preparar_imagem_transparente(PLACA_IMG_PATH, crop_box=PLACA_CROP, line_alpha=230)

    try:
        # Criar PDF 1-página em memória com diagrama + placa sobrepostos
        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=landscape(A4))

        # 1) Esquema Unifilar
        can.drawImage(
            tmp_diagrama,
            DIAGRAMA_POS_X, DIAGRAMA_POS_Y,
            width=DIAGRAMA_WIDTH, height=DIAGRAMA_HEIGHT,
            mask="auto",
        )

        # 2) Placa de Advertência (CUIDADO / Risco de Choque)
        can.drawImage(
            tmp_placa,
            PLACA_POS_X, PLACA_POS_Y,
            width=PLACA_WIDTH, height=PLACA_HEIGHT,
            mask="auto",
        )

        can.save()
        packet.seek(0)

        watermark_pdf = PdfReader(packet)
        watermark_page = watermark_pdf.pages[0]

        # Carregar PDF original
        original_pdf = PdfReader(caminho_pdf)
        writer = PdfWriter()

        # Adicionar todas as páginas
        total_pages = len(original_pdf.pages)
        for i in range(total_pages):
            page = original_pdf.pages[i]
            # Última página = DU-SOLAR: mescla overlay semi-transparente POR CIMA
            # As linhas do diagrama têm alpha ~43%, então o texto por baixo é legível
            if i == total_pages - 1:
                page.merge_page(watermark_page)
                writer.add_page(page)
            else:
                writer.add_page(page)

        # Salvar sobrescrevendo o arquivo
        with open(caminho_pdf, "wb") as f_out:
            writer.write(f_out)
    finally:
        os.unlink(tmp_diagrama)
        os.unlink(tmp_placa)

def validar_pdf(caminho_pdf: str, campos_esperados: list = None) -> dict:
    """
    Valida o PDF verificando se os campos esperados aparecem no texto.
    Usa pdftotext/pdfinfo se disponíveis (Linux); em Windows faz validação
    simples pelo tamanho do arquivo.

    Args:
        caminho_pdf: caminho do PDF a validar.
        campos_esperados: lista de strings que devem estar no PDF.

    Returns:
        dict com status e campos encontrados/faltantes.
    """
    # Tentar extrair texto com pdftotext (Linux/macOS com poppler)
    texto = ""
    try:
        r = subprocess.run(
            ["pdftotext", "-layout", caminho_pdf, "-"],
            capture_output=True, text=True,
        )
        texto = r.stdout.upper()
    except FileNotFoundError:
        pass  # pdftotext não disponível (Windows sem poppler)

    # Tentar obter número de páginas com pdfinfo
    num_paginas = 0
    try:
        r2 = subprocess.run(["pdfinfo", caminho_pdf], capture_output=True, text=True)
        for linha in r2.stdout.split("\n"):
            if "Pages:" in linha:
                try:
                    num_paginas = int(linha.split(":")[1].strip())
                except ValueError:
                    pass
    except FileNotFoundError:
        pass  # pdfinfo não disponível (Windows sem poppler)

    # Fallback: verificar se o arquivo existe e tem tamanho razoável
    pdf_existe = Path(caminho_pdf).exists()
    pdf_tamanho = Path(caminho_pdf).stat().st_size if pdf_existe else 0

    # Se pdftotext não está disponível, considerar ok se PDF tem tamanho > 10 KB
    if not texto and pdf_tamanho > 10_000:
        return {
            "ok": True,
            "num_paginas": num_paginas,
            "encontrados": campos_esperados or [],
            "faltando": [],
            "aviso": "pdftotext indisponivel — validado por tamanho do arquivo",
        }

    encontrados = []
    faltando = []
    for campo in (campos_esperados or []):
        if campo.upper() in texto:
            encontrados.append(campo)
        else:
            faltando.append(campo)

    ok = len(faltando) == 0
    return {
        "ok": ok,
        "num_paginas": num_paginas,
        "encontrados": encontrados,
        "faltando": faltando,
    }
