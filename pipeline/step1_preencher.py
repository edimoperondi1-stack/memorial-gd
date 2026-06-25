"""
step1_preencher.py
------------------
Recebe um DadosProjeto e preenche uma CÓPIA do template Excel.
NUNCA modifica o arquivo original.

Retorna o caminho do arquivo temporário preenchido.
"""

import shutil
import os
from pathlib import Path
from openpyxl import load_workbook
from modelos import DadosProjeto, sanitize_filename_part

TEMPLATE_PATH = Path(__file__).parent.parent / "MEMORIAL_GD_v4-22022022 (54).xlsx"


def preencher_template(dados: DadosProjeto, pasta_saida: str = None) -> str:
    """
    Cria uma cópia do template, preenche com os dados e retorna o caminho.

    Args:
        dados: DadosProjeto com todos os campos de entrada.
        pasta_saida: pasta onde salvar o arquivo temporário.
                     Se None, usa /tmp.

    Returns:
        Caminho absoluto do arquivo .xlsx preenchido.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template não encontrado: {TEMPLATE_PATH}")

    pasta = Path(pasta_saida) if pasta_saida else Path("/tmp")
    pasta.mkdir(parents=True, exist_ok=True)

    destino = pasta / f"_temp_{sanitize_filename_part(dados.codigo_uc)}.xlsx"
    shutil.copy2(TEMPLATE_PATH, destino)

    wb = load_workbook(str(destino))

    _preencher_solicitacao(wb, dados)
    _preencher_md_solar(wb, dados)
    _preencher_relacao_carga(wb, dados)
    _preencher_fsa(wb, dados)
    _preencher_formulario(wb, dados)

    wb.save(str(destino))
    wb.close()

    print(f"  [step1] Template preenchido: {destino}")
    return str(destino)


# ─────────────────────────────────────────────────────────────────────────────
# Funções internas de preenchimento por aba
# ─────────────────────────────────────────────────────────────────────────────

def _safe_write(ws, coord, value):
    try:
        ws[coord].value = value
    except AttributeError:
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(rng.min_row, rng.min_col).value = value
                break

def _preencher_solicitacao(wb, dados: DadosProjeto):
    """Preenche a aba SOLICITACAO."""
    ws = wb["SOLICITACAO"]

    ws["C7"]  = dados.codigo_uc
    ws["I7"]  = dados.classe
    ws["C8"]  = dados.titular
    ws["C9"]  = dados.logradouro
    ws["C10"] = dados.numero
    ws["E10"] = dados.bairro
    ws["I10"] = dados.uf
    ws["K10"] = dados.cep
    ws["I11"] = dados.cidade
    ws["C11"] = dados.email
    ws["C12"] = dados.telefone
    ws["I12"] = dados.celular
    ws["C13"] = dados.cpf_cnpj
    ws["D15"] = dados.potencia_instalada_kw
    ws["J15"] = dados.tensao_atendimento_v
    ws["D16"] = dados.tipo_conexao
    ws["D17"] = dados.tipo_ramal
    ws["D20"] = dados.tipo_fonte
    ws["H20"] = dados.tipo_geracao

    # Responsável técnico
    ws["D26"] = dados.resp_nome
    ws["C27"] = dados.resp_telefone
    ws["G27"] = dados.resp_email


def _preencher_md_solar(wb, dados: DadosProjeto):
    """Preenche a aba MD-SOLAR."""
    ws = wb["MD-SOLAR"]

    # Identificação da UC
    ws["C6"]  = dados.codigo_uc
    ws["G6"]  = dados.classe
    ws["J6"]  = dados.cpf_cnpj
    ws["C7"]  = dados.titular
    ws["C8"]  = dados.logradouro
    ws["C9"]  = dados.numero
    ws["E9"]  = dados.bairro
    ws["I9"]  = dados.cidade
    ws["C10"] = dados.email
    ws["I10"] = dados.uf
    ws["K10"] = dados.cep
    ws["C11"] = dados.telefone
    ws["G11"] = dados.celular
    # NÃO escrever J11/K11: são os RÓTULOS "Nº de fases:" e "Ramal". Os valores
    # vêm de fórmulas no template — J12 calcula o nº de fases a partir do tipo de
    # padrão (B13) e K12 = CONFIG!X2 (o ramal). Escrever J11/K11 sobrescrevia os
    # rótulos (saía "3"/"AÉREO" no lugar de "Nº de fases:"/"Ramal").

    # Padrão elétrico
    ws["B13"] = dados.tipo_padrao
    ws["D13"] = dados.nivel_tensao_v
    # Potência máx. disponibilizada: só preenche se informada; senão fica em
    # branco (a distribuidora define depois), igual ao formulário oficial.
    if dados.potencia_max_disponivel_kw:
        ws["G13"] = dados.potencia_max_disponivel_kw
    ws["B15"] = dados.disjuntor_geral_a
    ws["D15"] = dados.fator_potencia
    ws["G15"] = dados.demanda_contratada_kw
    # Proteção — valores vão na ROW 16 (row 14 = cabeçalhos, merged com row 15)
    # DU-SOLAR referencia I16, J16, K16, L16 para o diagrama unifilar
    ws["I16"] = dados.dps_ca_ka
    ws["J16"] = dados.disjuntor_ca_a
    ws["K16"] = dados.dps_cc_ka
    ws["L16"] = dados.disjuntor_cc_a

    # Modalidade e trafo — D16 é label "Potencia Trafo:", valor vai em D17
    ws["B17"] = dados.modalidade
    ws["D17"] = dados.potencia_trafo_kw
    ws["E17"] = dados.num_hastes

    # Coordenadas UTM
    ws["F19"] = dados.fuso
    ws["H19"] = dados.coord_x_long or None
    ws["J19"] = dados.coord_y_lat  or None

    # Configuração técnica
    ws["B22"] = dados.nivel_tensao_tipo
    ws["C22"] = dados.cabos_por_fase
    ws["D22"] = dados.potencia_geracao_kwp
    ws["E22"] = dados.bitola_fase_mm2
    ws["F22"] = dados.bitola_neutro_mm2
    ws["G22"] = dados.bitola_terra_mm2
    ws["H22"] = dados.gd_ja_instalado
    ws["I22"] = dados.previsao_mes
    ws["J22"] = dados.previsao_ano
    ws["K22"] = dados.zona

    # Observações
    if dados.observacoes:
        ws["C23"] = dados.observacoes

    # Trafo acoplamento / exclusivo (perguntas da página 4 / MD-SOLAR)
    ws["J58"] = dados.trafo_acoplamento
    if str(dados.trafo_acoplamento).strip().upper() == "SIM" and dados.potencia_autotrafo_kw:
        ws["J59"] = dados.potencia_autotrafo_kw
    ws["J61"] = dados.trafo_exclusivo
    if str(dados.trafo_exclusivo).strip().upper() == "SIM" and dados.potencia_trafo_exclusivo_kw:
        ws["J62"] = dados.potencia_trafo_exclusivo_kw

    # Painéis (até 10 linhas, começando na linha 31)
    for i, painel in enumerate(dados.paineis[:10]):
        row = 31 + i
        ws.cell(row=row, column=3).value  = painel.quantidade      # C
        ws.cell(row=row, column=4).value  = painel.fabricante       # D
        ws.cell(row=row, column=7).value  = painel.modelo           # G
        ws.cell(row=row, column=10).value = painel.area_m2          # J
        ws.cell(row=row, column=11).value = painel.potencia_kw      # K

    # Inversores (até 10 linhas, começando na linha 46)
    for i, inv in enumerate(dados.inversores[:10]):
        row = 46 + i
        ws.cell(row=row, column=3).value  = inv.quantidade          # C
        ws.cell(row=row, column=4).value  = inv.fabricante          # D
        ws.cell(row=row, column=7).value  = inv.modelo              # G
        ws.cell(row=row, column=10).value = inv.potencia_kw         # J
        ws.cell(row=row, column=12).value = inv.tensao_nominal_v    # L


def _preencher_relacao_carga(wb, dados: DadosProjeto):
    """Preenche a aba RELACAO DE CARGA (até 20 equipamentos, linha 16 em diante)."""
    ws = wb["RELACAO DE CARGA"]

    for i, item in enumerate(dados.carga_instalada[:20]):
        row = 16 + i
        qtd, equipamento, pot_w, fator = item
        ws.cell(row=row, column=2).value  = qtd          # B = quantidade
        ws.cell(row=row, column=3).value  = equipamento  # C = equipamento
        ws.cell(row=row, column=6).value  = pot_w        # F = potência unitária (W)
        ws.cell(row=row, column=8).value  = fator        # H = fator de demanda


def _preencher_fsa(wb, dados: DadosProjeto):
    """Preenche a aba FSA correspondente ao tipo_fsa."""
    aba = dados.tipo_fsa
    if aba not in wb.sheetnames:
        return

    ws = wb[aba]

    # 1. Identificação da UC
    ws["C7"]  = dados.codigo_uc
    ws["I7"]  = dados.classe
    ws["C8"]  = dados.titular
    ws["C9"]  = dados.logradouro
    ws["C10"] = dados.numero
    ws["E10"] = dados.bairro
    ws["I10"] = dados.uf
    ws["K10"] = dados.cep
    ws["I11"] = dados.cidade
    ws["C11"] = dados.email
    ws["C12"] = dados.telefone
    ws["I12"] = dados.celular
    ws["C13"] = dados.cpf_cnpj
    ws["D15"] = dados.potencia_instalada_kw
    ws["J15"] = dados.tensao_atendimento_v
    ws["D16"] = dados.tipo_conexao
    ws["D17"] = dados.tipo_ramal

    # 3. Dados da Geração
    ws["D20"] = dados.tipo_fonte        # Tipo da Fonte (ex: "SOLAR FOTOVOLTAICA")
    ws["H20"] = dados.tipo_geracao      # Tipo de Geração (H20:K20 merged, dentro do print area)

    # 4. Documentações a serem anexadas no AWGPE (checkmarks) — SOMENTE nos
    # formulários FSA MICRO (Solicitação de Acesso). No "Orçamento de Conexão"
    # (aba SOLICITACAO) as células K22/K23/K24 pertencem às MESCLAS dos CONTATOS
    # (D22:K22 Empresa, G23:K23 E-mail, G24:K24 LINK GISA); escrever "X" nelas
    # sobrescrevia o VLOOKUP dos contatos (saíam como "X"). O checklist do
    # Orçamento de Conexão fica na aba FORMULARIO (página 3).
    if aba != "SOLICITACAO":
        _safe_write(ws, "K22", "X")  # 1. Registro conselho profissional
        _safe_write(ws, "K23", "X")  # 2. Diagrama Unifilar
        _safe_write(ws, "K24", "X")  # 3. Certificado de conformidade
        # Item 4 (registro ANEEL) — não precisa
        if getattr(dados, "ucs_beneficiarias", False):
            _safe_write(ws, "K26", "X")  # 5. Lista de UCs participantes

    # 6. Dados do Solicitante (engenheiro responsável)
    _safe_write(ws, "D34", dados.resp_nome)
    _safe_write(ws, "C35", dados.resp_telefone)
    _safe_write(ws, "G35", dados.resp_email)


# ─── Mapeamento item → row no FORMULÁRIO ────────────────────────────────────
# Cada item é um checkbox; "X" marca presença, "SIM"/"NÃO" para 3.1
FORMULARIO_MAP = {
    "1":    6,
    "1.1":  7,
    "1.2":  8,
    "1.3":  9,
    "1.4":  10,
    "1.5":  11,
    "1.6":  12,
    "1.7":  13,
    "1.8":  14,
    "1.9":  15,
    "2.1":  17,
    "2.2":  18,
    "2.3":  19,
    "2.4":  20,
    "2.5":  21,
    "2.6":  22,
    "2.7":  23,
    "2.8":  24,
    "2.9":  25,
    "2.10": 26,
    "2.11": 27,
    "3.1":  29,
    "3.2":  30,
    "3.3":  31,
    "3.4":  32,
}

# Seleção padrão conforme prática comum
FORMULARIO_DEFAULTS = {
    "1":   "X",
    "1.2": "X",
    "1.3": "X",
    "1.4": "X",
    "2.1": "X",
    "2.3": "X",
    "2.4": "X",
    "2.5": "X",
    "3.1": "NÃO",
    "3.4": "X",
}


def _preencher_formulario(wb, dados: DadosProjeto):
    """Preenche a aba FORMULARIO com as marcações X nos itens selecionados."""
    if "FORMULARIO" not in wb.sheetnames:
        return

    ws = wb["FORMULARIO"]

    # Usar seleção customizada se fornecida, senão usar defaults
    items = getattr(dados, "formulario_items", None) or FORMULARIO_DEFAULTS

    for item_key, valor in items.items():
        row = FORMULARIO_MAP.get(item_key)
        if row and valor:
            coord = f"K{row}"
            _safe_write(ws, coord, valor)
