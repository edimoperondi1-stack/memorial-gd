"""
Verifica se um XLSX de saida atende aos 8 criterios de conformidade com a planilha oficial.
Uso: python pipeline/test_xlsx_correcoes.py <caminho_xlsx>
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def verificar(caminho):
    wb = load_workbook(caminho)
    ws = wb.active  # deve ser SAIDA (unica aba)
    resultados = []

    # 1. AT2:DR2 — string vazia, nao None
    nones = []
    for col in range(column_index_from_string("AT"), column_index_from_string("DR") + 1):
        val = ws.cell(row=2, column=col).value
        if val is None:
            from openpyxl.utils import get_column_letter
            nones.append(get_column_letter(col))
    resultados.append(("1-AT2:DR2 string vazia", len(nones) == 0, f"None em: {nones[:5]}..." if nones else ""))

    # 2. F2, G2, M2 numericos
    for col_letter in ["F", "G", "M"]:
        val = ws[f"{col_letter}2"].value
        is_num = isinstance(val, (int, float))
        resultados.append((f"2-{col_letter}2 numerico", is_num, f"tipo={type(val).__name__}, val={val}"))

    # 3. Sheet protection
    resultados.append(("3-Sheet protection", ws.protection.sheet == True, f"sheet={ws.protection.sheet}"))

    # 4. Sheet hidden
    resultados.append(("4-Sheet hidden", ws.sheet_state == "hidden", f"state={ws.sheet_state}"))

    # 5. AJ2 format General
    fmt = ws["AJ2"].number_format
    resultados.append(("5-AJ2 General", fmt == "General", f"format={fmt}"))

    # 6. DL/DM width 36.43
    from openpyxl.utils import get_column_letter
    for col_num, col_name in [(116, "DL"), (117, "DM")]:
        w = ws.column_dimensions[get_column_letter(col_num)].width
        ok = w is not None and abs(w - 36.43) < 0.01
        resultados.append((f"6-{col_name} width", ok, f"width={w}"))

    # 7. Header/footer margins 0.315
    for margin_name in ["header", "footer"]:
        val = getattr(ws.page_margins, margin_name)
        ok = val is not None and abs(val - 0.315) < 0.001
        resultados.append((f"7-{margin_name} margin", ok, f"margin={val}"))

    # 8. No fitToWidth/fitToHeight
    ftw = ws.page_setup.fitToWidth
    fth = ws.page_setup.fitToHeight
    ok_fit = (ftw is None or ftw == 0) and (fth is None or fth == 0)
    resultados.append(("8-No fitTo", ok_fit, f"fitToWidth={ftw}, fitToHeight={fth}"))

    wb.close()

    # Relatorio
    all_pass = True
    for nome, ok, detalhe in resultados:
        status = "PASS" if ok else "FAIL"
        if not ok:
            all_pass = False
        print(f"  [{status}] {nome}  {detalhe}")

    print(f"\n{'TODOS OK' if all_pass else 'FALHAS ENCONTRADAS'}")
    return all_pass


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python pipeline/test_xlsx_correcoes.py <caminho.xlsx>")
        sys.exit(1)
    ok = verificar(sys.argv[1])
    sys.exit(0 if ok else 1)
