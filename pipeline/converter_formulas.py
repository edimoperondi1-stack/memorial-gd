"""
converter_formulas.py
---------------------
Converte fórmulas em valores calculados diretamente no XML do xlsx (nível ZIP).

Isso é necessário porque:
  - O LibreOffice macro deleta abas antes de exportar PDF
  - Fórmulas que referenciam abas deletadas viram #REF!
  - Convertendo fórmulas → valores ANTES da exportação, os dados ficam preservados

A conversão lê os valores de openpyxl (data_only=True) e modifica o XML
da célula para trocar a fórmula por um valor constante.

Formato XML de uma célula com fórmula:
    <c r="A1" t="str"><f>='MD-SOLAR'!C6</f><v>3941140</v></c>

Formato XML de uma célula com valor:
    <c r="A1" t="s"><v>0</v></c>      (string shared)
    <c r="A1" t="str"><v>3941140</v></c>  (inline string)
    <c r="A1"><v>42</v></c>             (número)
"""

import re
import shutil
import tempfile
import zipfile
from pathlib import Path


def converter_formulas_para_valores(
    caminho_xlsx: str,
    caminho_template_ou_recalc: str,
    abas_converter: list[str],
) -> str:
    """
    Substitui fórmulas por valores nas abas especificadas.

    A estratégia é simples e robusta:
      1. Ler valores calculados com openpyxl data_only=True
      2. No XML do sheet, remover elementos <f>...</f> e garantir que <v> tem o valor certo

    Args:
        caminho_xlsx: arquivo xlsx a modificar (in-place).
        caminho_template_ou_recalc: arquivo com valores calculados (pode ser o mesmo).
        abas_converter: lista de nomes de abas para converter.

    Returns:
        caminho_xlsx (modificado in-place).
    """
    from openpyxl import load_workbook

    # 1. Ler valores calculados
    wb_val = load_workbook(caminho_template_ou_recalc, data_only=True)

    # 2. Mapear sheet name → sheet file no zip
    sheet_file_map = _mapear_sheets(caminho_xlsx)

    # 3. Coletar valores por sheet
    valores_por_sheet = {}
    for aba in abas_converter:
        if aba not in wb_val.sheetnames:
            continue
        ws = wb_val[aba]
        vals = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    vals[cell.coordinate] = cell.value
        valores_por_sheet[aba] = vals

    wb_val.close()

    # 4. Modificar XMLs no ZIP
    tmp_path = tempfile.mktemp(suffix=".xlsx")

    with zipfile.ZipFile(caminho_xlsx, "r") as z_in, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as z_out:

        for item in z_in.namelist():
            data = z_in.read(item)

            # Verificar se este item é um sheet que precisamos converter
            sheet_modificado = False
            for aba_nome, vals in valores_por_sheet.items():
                sheet_file = sheet_file_map.get(aba_nome)
                if sheet_file and item == f"xl/{sheet_file}":
                    xml_str = data.decode("utf-8")
                    xml_str = _substituir_formulas(xml_str, vals)
                    z_out.writestr(item, xml_str)
                    sheet_modificado = True
                    break

            if not sheet_modificado:
                z_out.writestr(item, data)

    shutil.move(tmp_path, caminho_xlsx)
    return caminho_xlsx


def _mapear_sheets(caminho_xlsx: str) -> dict:
    """Retorna dict de {sheet_name: 'worksheets/sheetN.xml'}."""
    with zipfile.ZipFile(caminho_xlsx) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    # rId → Target
    rid_map = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))

    # sheet name → rId
    sheets = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml)

    result = {}
    for name, rid in sheets:
        name_decoded = name.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
        target = rid_map.get(rid, "")
        result[name_decoded] = target

    return result


def _substituir_formulas(xml: str, valores: dict) -> str:
    """
    No XML de um sheet, para cada célula com fórmula (<f>...</f>),
    remove a fórmula e define o valor como constante.
    """
    def _replace_cell(match):
        cell_tag = match.group(0)

        # Se não tem fórmula, não mexer
        if "<f>" not in cell_tag and "<f " not in cell_tag:
            return cell_tag

        # Extrair coordenada (r="A1")
        r_match = re.search(r'r="([^"]+)"', cell_tag)
        if not r_match:
            return cell_tag
        coord = r_match.group(1)

        # Buscar valor calculado
        valor = valores.get(coord)
        if valor is None:
            # Sem valor → remover fórmula, deixar célula vazia
            cell_tag = re.sub(r"<f[^>]*>.*?</f>", "", cell_tag)
            return cell_tag

        # Remover a fórmula
        cell_tag = re.sub(r"<f[^>]*>.*?</f>", "", cell_tag)

        # Determinar tipo e valor
        if isinstance(valor, (int, float)):
            # Número: type não especificado ou "n", <v>numero</v>
            cell_tag = re.sub(r't="[^"]*"', '', cell_tag)  # remover type existente
            # Substituir ou inserir <v>
            if "<v>" in cell_tag:
                cell_tag = re.sub(r"<v>.*?</v>", f"<v>{valor}</v>", cell_tag)
            else:
                cell_tag = cell_tag.replace("</c>", f"<v>{valor}</v></c>")
        else:
            # String: type="inlineStr" com <is><t>texto</t></is>
            # Ou type="str" com <v>texto</v>
            str_valor = _escape_xml(str(valor))
            # Remover type e value existentes
            cell_tag = re.sub(r't="[^"]*"', 't="str"', cell_tag)
            if 't="str"' not in cell_tag:
                cell_tag = cell_tag.replace(f'r="{coord}"', f'r="{coord}" t="str"')
            # Substituir ou inserir <v>
            if "<v>" in cell_tag:
                cell_tag = re.sub(r"<v>.*?</v>", f"<v>{str_valor}</v>", cell_tag)
            else:
                cell_tag = cell_tag.replace("</c>", f"<v>{str_valor}</v></c>")

        return cell_tag

    # Regex para capturar tags <c ...>...</c> (células)
    return re.sub(r"<c\s[^>]*>.*?</c>", _replace_cell, xml, flags=re.DOTALL)


def _escape_xml(text: str) -> str:
    """Escapa caracteres especiais para XML."""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )
