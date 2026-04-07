"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from office.soffice import get_soffice_env

from openpyxl import load_workbook

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def _get_macro_dir():
    """Retorna o diretório de macros do LibreOffice para o sistema atual."""
    system = platform.system()
    if system == "Darwin":
        return os.path.expanduser(MACRO_DIR_MACOS)
    elif system == "Windows":
        appdata = os.environ.get("APPDATA", os.path.expanduser("~"))
        return os.path.join(appdata, "LibreOffice", "4", "user", "basic", "Standard")
    else:
        return os.path.expanduser(MACRO_DIR_LINUX)


def _get_soffice_cmd():
    """Retorna o comando soffice. O PATH já é configurado pelo get_soffice_env()."""
    if platform.system() == "Windows":
        # Verificar caminhos comuns como fallback
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for c in candidates:
            if os.path.exists(c):
                return c
        return "soffice.exe"
    return "soffice"


def has_gtimeout():
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def setup_libreoffice_macro():
    macro_dir = _get_macro_dir()
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text(encoding="utf-8")
    ):
        return True

    if not os.path.exists(macro_dir):
        soffice = _get_soffice_cmd()
        try:
            subprocess.run(
                [soffice, "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=30,
                env=get_soffice_env(),
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass  # soffice pode não estar no PATH, tentar criar o dir mesmo assim
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    soffice = _get_soffice_cmd()
    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    system = platform.system()
    if system == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif system == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd
    # Windows: sem timeout externo (usa timeout do subprocess.run)

    run_timeout = timeout + 30 if system == "Windows" else None

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        env=get_soffice_env(),
        timeout=run_timeout,
    )

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or result.stdout or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": f"LibreOffice macro not configured properly. stderr: {error_msg[:500]}"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result_data = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result_data["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result_data["total_formulas"] = formula_count

        return result_data

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
